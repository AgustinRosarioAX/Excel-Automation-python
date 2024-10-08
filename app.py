import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_excel(filename):

    wb= xl.load_workbook(filename)
    sheet = wb['Sheet1']
    #cell = sheet['a1']
    cell= sheet.cell(1,1)

    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        corrected_price= cell.value*0.9
        corrected_price_round2= round(corrected_price,2)
        corrected_price_cell= sheet.cell(row,4)
        corrected_price_cell.value= corrected_price_round2
    sheet['d1']= "new_price"
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)


#print(sheet.max_row)